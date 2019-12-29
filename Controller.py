from Session import Session
# include standard modules
import getopt, sys
from openpyxl import Workbook
from openpyxl import load_workbook
from NutritionInfo import NutritionInfo
import collections

# read commandline arguments, first
fullCmdArguments = sys.argv

# - further arguments
argumentList = fullCmdArguments[1:]

unixOptions = "t:a:s:f:"
gnuOptions = ["token=", "accountid=", "save=", "file="]
try:
    arguments, values = getopt.getopt(argumentList, unixOptions, gnuOptions)
except getopt.error as err:
    # output error, and return with an error code
    print (str(err))
    sys.exit(2)

# Declarations
token_ = None
accountid_ = None
write_file = None
path = None
tmp_file = None
# evaluate given options
for currentArgument, currentValue in arguments:
    if currentArgument in ("-t", "--token"):
        token_ = currentValue
    elif currentArgument in ("-a", "--accountid"):
        try:
            accountid_ = int(currentValue)
        except:
            print('Account id not a valid integer')
            sys.exit(2)
    elif currentArgument in ("-s", "--save"):
        path = currentValue
        try:
            open(path, 'w')
        except:
            print('Failed to write file: ' + str(path))
            sys.exit(2)
        if not path.endswith('.xlsx'):
            print('File must be a .xlsx')
            sys.exit(2)
    elif currentArgument in ("-f", "--file"):
        write_file = currentValue
        try:
            open(write_file, 'r')
        except:
            print('Failed to open file: ' + str(write_file))
        if not write_file.endswith('.xlsx'):
            print('File must be a .xlsx')
            sys.exit(2)

sess = Session(token=token_, accountid=accountid_)

before_write_foods = sess.getFoodList()
before_write_data = [sess.getFood(food).getAsRow() for food in before_write_foods]

cols = ['ID (If modifying existing)', 'Image (not supported)', 'Name', 'Serving Size', 'Serving Size measurement', 'Calories (Cal)', 'Kcal from Fat (Cal)', 'Total Fat (g)', 'Saturated Fat (g)', 'Trans Fat (g)', 'Polysaturated Fat (g)', 'Monounsaturated Fat (g)', 'Cholesterol (mg)', 'Sodium (mg)', 'Total Carbs (g)', 'Dietary Fiber (g)', 'Sugars (g)', 'Protein (g)', 'VD (mcg)', 'Calcium (mg)', 'Iron (mg)', 'Potassium (mg)', 'VA (mg)', 'VC (mg)']
# If we are not writing data back to scale/server
if write_file is None:
    wb = Workbook()
    ws = wb.active
    ws.append(cols)
    for row in before_write_data:
        ws.append(row)
    # Save the file
    wb.save(path)
else:
    wb = load_workbook(write_file)
    ws = wb[wb.sheetnames[0]]
    new_data = []
    for row in ws.iter_rows(min_row=2, max_col=len(cols)):
        row_ = []
        for cell in row:
            val = cell.value
            if val is None:
                val = -1
            row_.append(val)
        new = True
        for oldRow in before_write_data:
            if collections.Counter(oldRow) == collections.Counter(row_):
                new = False
        if new:
            new_data.append(row_)
    print('Found ' + str(len(new_data)) + ' new foods.')
    for data in new_data:
        if data[5] < 0:
            print('Calories must be a non-negative integer')

        n_data = {'foodID': data[0], 'foodname': data[2], 'image': data[1], 'weight': data[3], 'weight_unit': data[4], 'calories': data[5], 'calfromfat': data[6], 'totalfat': data[7], 'saturatedfat': data[8], 'transfat': data[9], 'fat_poly': data[10], 'fat_mono': data[11], 'cholesterol': data[12], 'sodium': data[13], 'totalcarbs': data[14], 'dietaryfiber': data[15], 'sugars': data[16], 'protein': data[17], 'vd': data[18], 'calcium': data[19], 'iron': data[20], 'potassium': data[21], 'va': data[22], 'vc': data[23]}
        if data[0] == -1:
            if sess.createFood(NutritionInfo(data[2], n_data)):
                print('Succesfully created food: ' + str(data[2]))
            else:
                print('Failed to create food: ' + str(data[2]))
        else:
            if sess.editFood(data[0], NutritionInfo(data[2], n_data)):
                print('Succesfully edited food: ' + str(data[2]))
            else:
                print('Failed to edit food: ' + str(data[2]))
